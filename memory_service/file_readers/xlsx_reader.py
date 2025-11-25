"""
XLSX and CSV file reader using openpyxl.
"""
from pathlib import Path
from typing import Optional
import openpyxl
import csv


def read_xlsx(path: Path) -> Optional[str]:
    """
    Extract text from an XLSX file.
    
    Args:
        path: Path to the XLSX file
        
    Returns:
        Extracted text as tabular representation, or None if extraction fails
    """
    try:
        workbook = openpyxl.load_workbook(str(path), data_only=True)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = [f"Sheet: {sheet_name}"]
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    sheet_text.append(row_text)
            
            if len(sheet_text) > 1:  # More than just the sheet name
                text_parts.append("\n".join(sheet_text))
        
        return "\n\n".join(text_parts)
    except Exception as e:
        print(f"Error reading XLSX {path}: {e}")
        return None


def read_csv(path: Path) -> Optional[str]:
    """
    Extract text from a CSV file.
    
    Args:
        path: Path to the CSV file
        
    Returns:
        Extracted text as tabular representation, or None if extraction fails
    """
    try:
        text_parts = []
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for row in reader:
                row_text = " | ".join(cell for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        print(f"Error reading CSV {path}: {e}")
        return None

